"""
    fileio.py
    Author : Milan Marocchi

    Reading and writing records
"""

import json
from json import JSONEncoder
from typing import Optional, Tuple
import numpy as np
import pandas as pd
import scipy.io as sio
from openpyxl import Workbook, load_workbook
import os, csv, io, time, math, shutil, tempfile
import numpy as np

def get_ticking_channel_map(collection: int) -> dict:
    if collection == 1:
        channel_map = {
            '1': 1,
            '2': 2,
            '3': 4,
            '4': 5,
            '5': 6,
            '6': 7,
            '7': 8, 
            '8': 11, #ECG1
            '9': 3,  #PPG
        }
    elif collection == 2:
        channel_map = {
            '1': 2,
            '2': 4,
            '3': 5,
            '4': 6,
            '5': 8,
            '6': 7,
            '7': 9,
            '8': 11, #ECG1
            '9': 1,  #PPG
            '10': 3, #ECG2
        }
    # This collection represents the processed vest_data channel format
    elif collection == -1:
        channel_map = {
            '1': 1,
            '2': 2,
            '3': 3,
            '4': 4,
            '5': 5,
            '6': 6,
            '7': 7,
            '8': 8, #ECG1
            '9': 9, #PPG
            '10': 10, #ECG2
        }
    else:
        raise ValueError(f"Incorrect collection number: {collection=}")

    return channel_map

def read_ticking_PCG(filename: str, channel: int, noise_mic: bool = False, collection: int = 1, max_len: Optional[int] = None) -> Tuple[np.ndarray, int]:
    """Read in the PCG from the ticking heart data."""
    channels = get_ticking_channel_map(collection)
    # filename += ".wav"
    signal, fs = read_signal_wav(filename)
    wav_channel = channels[str(channel)]
    wav_channel = wav_channel + 7 if noise_mic else wav_channel - 1 # -1 for MATLAB indexs, otherwise would be + 8 for NM

    return signal[:, wav_channel], fs


def read_ticking_ECG(filename: str, collection: int = 1, channel: int = 1, max_len: Optional[int] = None) -> Tuple[np.ndarray, int]:
    """Read in the ECG from the ticking heart data."""
    ecg_channel = 'E' if channel == 1 else 'E2'
    ecg = get_ticking_channel_map(collection)[ecg_channel] - 1 # -1 for MATLAB indexs
    # filename += ".wav"
    signal, fs = read_signal_wav(filename)

    return signal[:, ecg], fs

def read_ticking_PPG(filename: str, collection: int = 1, max_len: Optional[int] = None) -> Tuple[np.ndarray, int]:
    """Read in the ECG from the ticking heart data."""
    ppg_channel = 'P'
    ppg = get_ticking_channel_map(collection)[ppg_channel] - 1 # -1 for MATLAB indexs
    # filename += ".wav"
    signal, fs = read_signal_wav(filename)

    return signal[:, ppg], fs

def read_signal_wav(filename: str) -> Tuple[np.ndarray, int]:
    """
    Reads in a signal from a wav file then converts it into the same format that matlab would output.
    Outputs the sampling freq as well as the signal
    """
    if ".wav" not in filename:
        filename += ".wav"

    Fs, signal = sio.wavfile.read(filename)

    if signal.dtype == np.int16:
        max_val = np.iinfo(np.int16).max
    elif signal.dtype == np.int32:
        max_val = np.iinfo(np.int32).max
    elif signal.dtype == np.int64:
        max_val = np.iinfo(np.int64).max
    elif signal.dtype == np.float32 or signal.dtype == np.float64:
        # print('matt')
        # input(signal.dtype)
        return signal.astype(np.float32), Fs
    else:
        raise ValueError("Unsupported data type")

    # Convert to float 32
    signal = (signal / max_val).astype(np.float32)

    return signal, Fs


def save_signal_wav(signal: np.ndarray, fs: int, path: str):
    """
    Saves a signal as a wav file to the specified path.
    """
    if ".wav" not in path:
        path += ".wav"

    sio.wavfile.write(path, fs, signal)

def create_multi_wav(pcg_multi, c):
    # First 7 channels are heart facing
    # Last 7 channels are background mic facing
    ticking_wav = np.zeros((len(pcg_multi[0]), c))

    for i in range(0,c):
    # Create massive wav file so it is in the same format.
        ticking_wav[:, i] = pcg_multi[i]

    return ticking_wav

def save_ticking_signals(ticking_wav, fs, path):
    # Save the signals to the destination
    save_signal_wav(ticking_wav.astype(np.float32), fs, path)
    # Add signals to reference csv
    # patient = path.split("/")[-1]
    # with open(REFERENCE, "a") as fp:
    #     fp.write(f"{patient},{label}\n")

class NumpyEncoder(JSONEncoder):
    """
    Class to encode numpy data to a list to be stored in a json file.
    """

    def default(self, o: np.ndarray | list):
        if isinstance(o, np.ndarray):
            return o.tolist()
        return JSONEncoder.default(self, o)


def write_json_numpy(data: str, filepath: str):
    """
    Writes to a json using the numpy encoder
    """
    # NOTE: Add error checking so format is enforced.

    with open(filepath, "w") as out_file:
        json.dump(data, out_file, indent = 4)
def append_json_numpy(data: str, filepath: str):
    """
    Writes to a json using the numpy encoder
    """
    # NOTE: Add error checking so format is enforced.

    with open(filepath, "a") as out_file:
        json.dump(data, out_file, indent = 4)


def read_json_numpy(filepath: str) -> dict:
    """
    Reads from a json file and decodes the array to a numpy array.
    Excepts the format that is used when written
    """
    with open(filepath, "r") as in_file:
        json_data = json.load(in_file)

    return json_data

def save_val_metrics(filename, sheet_name, version_and_array):
    """
    Save [ver, np.array] into an Excel file.
    If file doesn't exist -> create it.
    If sheet doesn't exist -> create sheet.
    If sheet exists -> append a row.
    """
    ver, arr = version_and_array
    arr = np.array(arr).flatten().tolist()  # flatten in case it's >1D
    
    # Try to open existing file, else create new workbook
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        # remove default "Sheet" if it's still there
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            std = wb["Sheet"]
            wb.remove(std)

    # Use existing sheet or create new one
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        # Write header if new sheet
        header = ["version"] + [f"metric_{i}" for i in range(len(arr))]
        ws.append(header)

    # Append the new row
    row = [ver] + arr
    ws.append(row)

    # Save workbook
    wb.save(filename)


# --------- small helpers (local to this module) ---------
def _clean_number(x):
    try:
        if x is None: return None
        if isinstance(x, float) and (math.isnan(x) or math.isinf(x)):
            return None
        return float(x)
    except Exception:
        return None

def _atomic_write_bytes(path: str, data: bytes):
    """Write bytes safely: temp in same dir + fsync + atomic replace + fsync dir."""
    d = os.path.dirname(path) or "."
    os.makedirs(d, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=d, prefix=".tmp_", suffix=os.path.splitext(path)[1] or ".tmp")
    try:
        with os.fdopen(fd, "wb", closefd=True) as f:
            f.write(data)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp, path)
        # fsync the directory entry
        try:
            dirfd = os.open(d, os.O_DIRECTORY)
            try: os.fsync(dirfd)
            finally: os.close(dirfd)
        except Exception:
            pass
    finally:
        if os.path.exists(tmp):
            try: os.remove(tmp)
            except Exception: pass

def _write_csv_atomic(path: str, headers, rows):
    """Create CSV text from headers + rows (list-of-sequences) and write atomically."""
    sio = io.StringIO(newline="")
    w = csv.writer(sio)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    _atomic_write_bytes(path, sio.getvalue().encode("utf-8"))

def _load_or_new_workbook(excel_file: str):
    """Load workbook; if corrupt, back it up and return a fresh Workbook."""
    if os.path.exists(excel_file):
        try:
            return load_workbook(excel_file)
        except Exception:
            backup = excel_file.replace(".xlsx", f".corrupt_{int(time.time())}.xlsx")
            try: shutil.move(excel_file, backup)
            except Exception: pass
            wb = Workbook()
            return wb
    else:
        return Workbook()

def _atomic_save_workbook(wb, final_path: str):
    """Crash-safe workbook save: temp file + fsync + atomic replace + fsync dir."""
    out_dir = os.path.dirname(final_path) or "."
    os.makedirs(out_dir, exist_ok=True)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=out_dir) as tmp:
        tmp_path = tmp.name
    try:
        wb.save(tmp_path)
        # fsync temp
        try:
            with open(tmp_path, "rb") as f:
                f.flush()
                os.fsync(f.fileno())
        except Exception:
            pass
        os.replace(tmp_path, final_path)
        # fsync dir
        try:
            dirfd = os.open(out_dir, os.O_DIRECTORY)
            try: os.fsync(dirfd)
            finally: os.close(dirfd)
        except Exception:
            pass
    finally:
        if os.path.exists(tmp_path):
            try: os.remove(tmp_path)
            except Exception: pass

# --------- your rewritten function ---------
def average_results_and_save(parent_dir: str, numfolds: int, channels: list, ver: int):
    """
    Goes through all the channel files and averages the results.
    Saves averages into a text file, writes fold results to CSV (atomic),
    and appends results to an Excel file (atomic).
    Metrics (except MCC) are scaled by 100 and presented with 4 significant figures.
    """
    metrics = ['accuracy', 'sensitivity', 'specificity', 'MCC']  # Update these as needed
    scale_metrics = {'accuracy', 'sensitivity', 'specificity'}   # Metrics to scale

    excel_dir = os.path.dirname(parent_dir)  # directory above `parent_dir`
    excel_file = os.path.join(excel_dir, 'results_summary_single.xlsx')
    excel_file_val = os.path.join(excel_dir, 'VALresults_summary_single.xlsx')

    # Process each channel independently (do not mutate `chan`)
    for chan in channels:
        frag_sums = {m: 0.0 for m in metrics}
        sub_sums = {m: 0.0 for m in metrics}
        sub_sums_soft = {m: 0.0 for m in metrics}
        file_count = 0
        fold_results = []      # per-fold dicts for Excel/CSV
        loaded_val_metrics = []  # list of np arrays

        # ---- gather & compute per-fold ----
        for fold in range(1, numfolds + 1):
            res_path = os.path.join(parent_dir, f'ch{chan}', f'fold{fold}', 'results.txt')
            val_path = os.path.join(parent_dir, f'ch{chan}', f'fold{fold}', 'best_val_metrics.npy')

            data = read_json_numpy(res_path)  # raise if missing (keeps your behavior)
            fold_row = {'fold': fold, 'channel': chan}

            for m in metrics:
                frag_value = float(data['results_frag'][m])
                sub_value = float(data['results_sub'][m])
                sub_soft_value = float(data['results_sub_soft'][m])

                if m in scale_metrics:
                    frag_value *= 100.0
                    sub_value *= 100.0
                    sub_soft_value *= 100.0

                frag_value = _clean_number(round(frag_value, 4))
                sub_value = _clean_number(round(sub_value, 4))
                sub_soft_value = _clean_number(round(sub_soft_value, 4))

                frag_sums[m] += (frag_value or 0.0)
                sub_sums[m] += (sub_value or 0.0)
                sub_sums_soft[m] += (sub_soft_value or 0.0)

                fold_row[f'frag_{m}'] = frag_value
                fold_row[f'sub_{m}'] = sub_value
                fold_row[f'sub_soft_{m}'] = sub_soft_value

            fold_results.append(fold_row)
            file_count += 1

            loaded_val_metrics.append(np.load(val_path))

        # ---- compute averages ----
        if file_count == 0:
            # nothing found for this channel; skip gracefully
            continue

        frag_avg = {m: _clean_number(round(frag_sums[m] / file_count, 4)) for m in metrics}
        sub_avg =  {m: _clean_number(round(sub_sums[m]  / file_count, 4)) for m in metrics}
        sub_avg_soft = {m: _clean_number(round(sub_sums_soft[m] / file_count, 4)) for m in metrics}

        # ---- update res_single.txt (unchanged behavior) ----
        out_txt = os.path.join(parent_dir, 'res_single.txt')
        output_data = {
            f'ch{chan}_average_results_frag': frag_avg,
            f'ch{chan}_average_results_sub': sub_avg,
            f'ch{chan}_average_results_sub_soft': sub_avg_soft
        }
        if os.path.exists(out_txt):
            append_json_numpy(output_data, out_txt)
        else:
            write_json_numpy(output_data, out_txt)

        # ---- write per-channel CSV (atomic) ----
        csv_headers = (
            ['fold', 'channel'] +
            [f'frag_{m}' for m in metrics] + [''] +
            [f'sub_{m}' for m in metrics] + [''] +
            [f'sub_soft_{m}' for m in metrics]
        )
        csv_rows = []
        for r in fold_results:
            row = [
                r['fold'], r['channel'],
                *[r[f'frag_{m}'] for m in metrics], '',
                *[r[f'sub_{m}'] for m in metrics], '',
                *[r[f'sub_soft_{m}'] for m in metrics]
            ]
            # ensure floats are Excel-friendly (None ok)
            row = [_clean_number(x) if isinstance(x, float) else x for x in row]
            csv_rows.append(row)

        csv_file = os.path.join(parent_dir, f'ch{chan}_fold_results.csv')
        _write_csv_atomic(csv_file, csv_headers, csv_rows)

        # ---- append to Excel (atomic) ----
        wb = _load_or_new_workbook(excel_file)

        # remove default empty sheet if it's the only one
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1 and wb.active.max_row <= 1:
            try:
                wb.remove(wb['Sheet'])
            except Exception:
                pass

        sheet_name = f'ch{chan}'
        if sheet_name not in wb.sheetnames:
            sh = wb.create_sheet(title=sheet_name)
            sh.append(['Version', 'Fold', 'Channel'] +
                      [f'frag_{m}' for m in metrics] + [''] +
                      [f'sub_{m}' for m in metrics] + [''] +
                      [f'sub_soft_{m}' for m in metrics])
        else:
            sh = wb[sheet_name]
            # ensure headers if the first row is empty
            if sh.max_row == 1 and all(c.value is None for c in sh[1]):
                sh.append(['Version', 'Fold', 'Channel'] +
                          [f'frag_{m}' for m in metrics] + [''] +
                          [f'sub_{m}' for m in metrics] + [''] +
                          [f'sub_soft_{m}' for m in metrics])

        # Version marker
        sh.append([f'Version: {ver}'] + [''] * (len(sh[1]) - 1))

        # Fold rows
        for r in fold_results:
            row = [ver, r['fold'], r['channel']] + \
                  [r[f'frag_{m}'] for m in metrics] + [''] + \
                  [r[f'sub_{m}']  for m in metrics] + [''] + \
                  [r[f'sub_soft_{m}'] for m in metrics]
            row = [_clean_number(x) if isinstance(x, float) else x for x in row]
            sh.append(row)

        # Averages row
        avg_row = ['AVERAGE', '', ''] + \
                  [frag_avg[m] for m in metrics] + [''] + \
                  [sub_avg[m]  for m in metrics] + [''] + \
                  [sub_avg_soft[m] for m in metrics]
        avg_row = [_clean_number(x) if isinstance(x, float) else x for x in avg_row]
        sh.append(avg_row)

        _atomic_save_workbook(wb, excel_file)
        try: wb.close()
        except Exception: pass

        # ---- save VAL averages (kept as your original call) ----
        avg_val = np.mean(np.stack(loaded_val_metrics, axis=0), axis=0)
        save_val_metrics(excel_file_val, f'ch{chan}', [ver, avg_val])


# def average_results_and_save(parent_dir: str, numfolds: int, channels: list, ver: int):
#     """
#     Goes through all the channel files and averages the results.
#     Saves averages into a text file, writes fold results to separate CSV files,
#     and appends results to an Excel file with each channel on a separate sheet.
#     Metrics (except MCC) are scaled by 100 and presented with 4 significant figures.
#     """
#     metrics = ['accuracy', 'sensitivity', 'specificity', 'MCC']  # Update these as needed
#     scale_metrics = ['accuracy', 'sensitivity', 'specificity']  # Metrics to scale
#     excel_dir = os.path.dirname(parent_dir)  # Navigate to the directory above `parent_dir`
#     excel_file = os.path.join(excel_dir, 'results_summary_single.xlsx')
#     excel_file_val = os.path.join(excel_dir, 'VALresults_summary_single.xlsx')

#     # again = True
#     # chan = channels[0]
#     for chan in channels:
#         frag_sums = {metric: 0 for metric in metrics}
#         sub_sums = {metric: 0 for metric in metrics}
#         sub_sums_soft = {metric: 0 for metric in metrics}
#         file_count = 0
#         fold_results = []  # To store fold-wise results for CSV output
#         loaded_val_metrics=[]
#         for fold in range(1, numfolds + 1):
#             file = os.path.join(parent_dir, f'ch{chan}', f'fold{fold}', 'results.txt')
#             data = read_json_numpy(file)
#             fold_data = {'fold': fold, 'channel': chan}

#             # Accumulate values for results_frag, results_sub, and results_sub_soft
#             for metric in metrics:
#                 frag_value = data['results_frag'][metric]
#                 sub_value = data['results_sub'][metric]
#                 sub_soft_value = data['results_sub_soft'][metric]

#                 # Scale metrics if required
#                 if metric in scale_metrics:
#                     frag_value *= 100
#                     sub_value *= 100
#                     sub_soft_value *= 100

#                 # Format to 4 significant figures
#                 frag_value = round(frag_value, 4)
#                 sub_value = round(sub_value, 4)
#                 sub_soft_value = round(sub_soft_value, 4)

#                 frag_sums[metric] += frag_value
#                 sub_sums[metric] += sub_value
#                 sub_sums_soft[metric] += sub_soft_value

#                 # Add fold data for Excel and CSV
#                 fold_data[f'frag_{metric}'] = frag_value
#                 fold_data[f'sub_{metric}'] = sub_value
#                 fold_data[f'sub_soft_{metric}'] = sub_soft_value

#             fold_results.append(fold_data)
#             file_count += 1
#             val_file = os.path.join(parent_dir, f'ch{chan}', f'fold{fold}', 'best_val_metrics.npy')
#             loaded_val_metrics.append(np.load(val_file))

#         # Compute averages
#         frag_avg = {metric: round(frag_sums[metric] / file_count, 4) for metric in metrics}
#         sub_avg = {metric: round(sub_sums[metric] / file_count, 4) for metric in metrics}
#         sub_avg_soft = {metric: round(sub_sums_soft[metric] / file_count, 4) for metric in metrics}

#         # Prepare the output content
#         output_data = {
#             f'ch{chan}_average_results_frag': frag_avg,
#             f'ch{chan}_average_results_sub': sub_avg,
#             f'ch{chan}_average_results_sub_soft': sub_avg_soft
#         }

#         file_path = parent_dir + '/res_single.txt'
#         if os.path.exists(file_path):
#             append_json_numpy(output_data, file_path)
#         else:
#             write_json_numpy(output_data, file_path)

#         # Write fold results to CSV
#         csv_file = os.path.join(parent_dir, f'ch{chan}_fold_results.csv')
#         with open(csv_file, mode='w', newline='') as f:
#             writer = csv.DictWriter(f, fieldnames=['fold', 'channel'] + 
#                                     [f'frag_{metric}' for metric in metrics] + [''] + 
#                                     [f'sub_{metric}' for metric in metrics] + [''] + 
#                                     [f'sub_soft_{metric}' for metric in metrics])
#             writer.writeheader()
#             writer.writerows(fold_results)

#         # Append to Excel file in directory above `parent_dir`
#         if os.path.exists(excel_file):
#             workbook = load_workbook(excel_file)
#         else:
#             os.makedirs(excel_dir, exist_ok=True)  # Ensure the directory exists
#             workbook = Workbook()

#         if f'ch{chan}' not in workbook.sheetnames:
#             sheet = workbook.create_sheet(title=f'ch{chan}')
#         else:
#             sheet = workbook[f'ch{chan}']

#         # If the sheet is newly created, write headers
#         if sheet.max_row == 1:
#             sheet.append(['Version', 'Fold', 'Channel'] +
#                          [f'frag_{metric}' for metric in metrics] + [''] +
#                          [f'sub_{metric}' for metric in metrics] + [''] +
#                          [f'sub_soft_{metric}' for metric in metrics])
        
#         # Append a version marker and results
#         sheet.append([f'Version: {ver}'] + [''] * (len(sheet[1]) - 1))  # Add version marker
#         for result in fold_results:
#             row = [ver, result['fold'], result['channel']] + \
#                   [result[f'frag_{metric}'] for metric in metrics] + [''] + \
#                   [result[f'sub_{metric}'] for metric in metrics] + [''] + \
#                   [result[f'sub_soft_{metric}'] for metric in metrics]
#             sheet.append(row)

#         # Append averages row after all folds
#         avg_row = ['AVERAGE', ''] + \
#                   [''] + [frag_avg[metric] for metric in metrics] + [''] + \
#                   [sub_avg[metric] for metric in metrics] + [''] + \
#                   [sub_avg_soft[metric] for metric in metrics]
#         sheet.append(avg_row)

#         workbook.save(excel_file)

#         avg_val = np.mean(np.stack(loaded_val_metrics, axis=0), axis=0)
#         save_val_metrics(excel_file_val,f'ch{chan}',[ver,avg_val])

#         chan += 1
#         if not os.path.exists(os.path.join(parent_dir, f'ch{chan}')):
#             again = False

def average_results_and_save_multi(parent_dir: str, numfolds: int, chan_comb: str, ver: int):
    """
    Goes through all the channel combination files and averages the results.
    Saves averages into a text file, writes fold results to separate CSV files,
    and appends results to an Excel file with a sheet named after `chan_comb`.
    Metrics (except MCC) are scaled by 100 and presented with 4 significant figures.
    """
    metrics = ['accuracy', 'sensitivity', 'specificity', 'MCC']  # Update these as needed
    scale_metrics = ['accuracy', 'sensitivity', 'specificity']  # Metrics to scale
    excel_dir = os.path.dirname(parent_dir)  # Navigate to the directory above `parent_dir`
    excel_file = os.path.join(excel_dir, 'results_summary_multi.xlsx')
    excel_file_val = os.path.join(excel_dir, 'VALresults_summary_multi.xlsx')

    frag_sums = {metric: 0 for metric in metrics}
    sub_sums = {metric: 0 for metric in metrics}
    sub_sums_soft = {metric: 0 for metric in metrics}
    file_count = 0
    fold_results = []  # To store fold-wise results for CSV output
    loaded_val_metrics = []

    for fold in range(1, numfolds + 1):
        file = os.path.join(parent_dir, chan_comb, f'_MultiChan_fold{fold}', 'results.txt')
        data = read_json_numpy(file)
        fold_data = {'fold': fold, 'chan_comb': chan_comb}

        # Accumulate values for results_frag, results_sub, and results_sub_soft
        for metric in metrics:
            frag_value = data['results_frag'][metric]
            sub_value = data['results_sub'][metric]
            sub_soft_value = data['results_sub_soft'][metric]

            # Scale metrics if required
            if metric in scale_metrics:
                frag_value *= 100
                sub_value *= 100
                sub_soft_value *= 100

            # Format to 4 significant figures
            frag_value = round(frag_value, 4)
            sub_value = round(sub_value, 4)
            sub_soft_value = round(sub_soft_value, 4)

            frag_sums[metric] += frag_value
            sub_sums[metric] += sub_value
            sub_sums_soft[metric] += sub_soft_value

            # Add fold data for Excel and CSV
            fold_data[f'frag_{metric}'] = frag_value
            fold_data[f'sub_{metric}'] = sub_value
            fold_data[f'sub_soft_{metric}'] = sub_soft_value

        fold_results.append(fold_data)
        file_count += 1

        val_file = os.path.join(parent_dir, chan_comb, f'_MultiChan_fold{fold}', 'best_val_metrics.npy')
        loaded_val_metrics.append(np.load(val_file))

    # Compute averages
    frag_avg = {metric: round(frag_sums[metric] / file_count, 4) for metric in metrics}
    sub_avg = {metric: round(sub_sums[metric] / file_count, 4) for metric in metrics}
    sub_avg_soft = {metric: round(sub_sums_soft[metric] / file_count, 4) for metric in metrics}

    # Prepare the output content
    output_data = {
        chan_comb + '_average_results_frag': frag_avg,
        chan_comb + '_average_results_sub': sub_avg,
        chan_comb + '_average_results_sub_soft': sub_avg_soft
    }

    # Write results to res_multi.txt
    file_path = os.path.join(parent_dir, 'res_multi.txt')
    if os.path.exists(file_path):
        append_json_numpy(output_data, file_path)
    else:
        write_json_numpy(output_data, file_path)

    # Write fold results to CSV
    csv_file = os.path.join(parent_dir, f'{chan_comb}_fold_results.csv')
    with open(csv_file, mode='w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['fold', 'chan_comb'] + 
                                [f'frag_{metric}' for metric in metrics] + [''] + 
                                [f'sub_{metric}' for metric in metrics] + [''] + 
                                [f'sub_soft_{metric}' for metric in metrics])
        writer.writeheader()
        writer.writerows(fold_results)

    # Append to Excel file in directory above `parent_dir`
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
    else:
        os.makedirs(excel_dir, exist_ok=True)  # Ensure the directory exists
        workbook = Workbook()

    if chan_comb not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=chan_comb)
    else:
        sheet = workbook[chan_comb]

    # If the sheet is newly created, write headers
    if sheet.max_row == 1:
        sheet.append(['Version', 'Fold', 'Channel Combination'] +
                     [f'frag_{metric}' for metric in metrics] + [''] +
                     [f'sub_{metric}' for metric in metrics] + [''] +
                     [f'sub_soft_{metric}' for metric in metrics])
    
    # Append a Version marker and results
    sheet.append([f'Version: {ver}'] + [''] * (len(sheet[1]) - 1))  # Add Version marker
    for result in fold_results:
        row = [ver, result['fold'], result['chan_comb']] + \
              [result[f'frag_{metric}'] for metric in metrics] + [''] + \
              [result[f'sub_{metric}'] for metric in metrics] + [''] + \
              [result[f'sub_soft_{metric}'] for metric in metrics]
        sheet.append(row)

    # Append averages row after all folds
    avg_row = ['AVERAGE', '', ''] + \
              [frag_avg[metric] for metric in metrics] + [''] + \
              [sub_avg[metric] for metric in metrics] + [''] + \
              [sub_avg_soft[metric] for metric in metrics]
    sheet.append(avg_row)

    workbook.save(excel_file)

    avg_val = np.mean(np.stack(loaded_val_metrics, axis=0), axis=0)
    save_val_metrics(excel_file_val,chan_comb,[ver,avg_val])


def average_results_and_save_multi_svm(parent_dir: str, numfolds: int, chan_comb: str, svm_act: int, ver: int):
    """
    Goes through all the channel combination files and averages the results for SVM-based evaluation.
    Saves averages into a text file, writes fold results to separate CSV files,
    and appends results to an Excel file with a sheet named after `chan_comb`.
    Metrics (except MCC) are scaled by 100 and presented with 4 significant figures.
    """
    metrics = ['accuracy', 'sensitivity', 'specificity', 'MCC']  # Update these as needed
    scale_metrics = ['accuracy', 'sensitivity', 'specificity']  # Metrics to scale
    excel_dir = os.path.dirname(parent_dir)  # Navigate to the directory above `parent_dir`
    excel_file = os.path.join(excel_dir, 'results_summary_multi_svm.xlsx')

    frag_sums = {metric: 0 for metric in metrics}
    sub_sums = {metric: 0 for metric in metrics}
    file_count = 0
    fold_results = []  # To store fold-wise results for CSV output

    for fold in range(1, numfolds + 1):
        file = os.path.join(parent_dir, chan_comb, f'_MultiChan_fold{fold}', f'svm{svm_act}.txt')
        data = read_json_numpy(file)
        fold_data = {'fold': fold, 'chan_comb': chan_comb, 'svm_act': svm_act}

        # Accumulate values for results_frag and results_sub
        for metric in metrics:
            frag_value = data['results_frag'][metric]
            sub_value = data['results_sub'][metric]

            # Scale metrics if required
            if metric in scale_metrics:
                frag_value *= 100
                sub_value *= 100

            # Format to 4 significant figures
            frag_value = round(frag_value, 4)
            sub_value = round(sub_value, 4)

            frag_sums[metric] += frag_value
            sub_sums[metric] += sub_value

            # Add fold data for Excel and CSV
            fold_data[f'frag_{metric}'] = frag_value
            fold_data[f'sub_{metric}'] = sub_value

        fold_results.append(fold_data)
        file_count += 1

    # Compute averages
    frag_avg = {metric: round(frag_sums[metric] / file_count, 4) for metric in metrics}
    sub_avg = {metric: round(sub_sums[metric] / file_count, 4) for metric in metrics}

    # Prepare the output content
    output_data = {
        chan_comb + '_average_results_frag': frag_avg,
        chan_comb + '_average_results_sub': sub_avg
    }

    # Write results to res_multi_svm.txt
    file_path = os.path.join(parent_dir, 'res_multi_svm.txt')
    if os.path.exists(file_path):
        append_json_numpy(output_data, file_path)
    else:
        write_json_numpy(output_data, file_path)

    # Write fold results to CSV
    csv_file = os.path.join(parent_dir, f'{chan_comb}_svm{svm_act}_fold_results.csv')
    with open(csv_file, mode='w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['fold', 'chan_comb', 'svm_act'] +
                                [f'frag_{metric}' for metric in metrics] + [''] +
                                [f'sub_{metric}' for metric in metrics])
        writer.writeheader()
        writer.writerows(fold_results)

    # Append to Excel file in directory above `parent_dir`
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
    else:
        os.makedirs(excel_dir, exist_ok=True)  # Ensure the directory exists
        workbook = Workbook()

    if chan_comb not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=chan_comb)
    else:
        sheet = workbook[chan_comb]

    # If the sheet is newly created, write headers
    if sheet.max_row == 1:
        sheet.append(['Version', 'Fold', 'Channel Combination', 'SVM Activation'] +
                     [f'frag_{metric}' for metric in metrics] + [''] +
                     [f'sub_{metric}' for metric in metrics])
    
    # Append a version marker and results
    sheet.append([f'Version: {ver}'] + [''] * (len(sheet[1]) - 1))  # Add version marker
    for result in fold_results:
        row = [ver, result['fold'], result['chan_comb'], result['svm_act']] + \
              [result[f'frag_{metric}'] for metric in metrics] + [''] + \
              [result[f'sub_{metric}'] for metric in metrics]
        sheet.append(row)

    # Append averages row after all folds
    avg_row = ['AVERAGE', '', '', ''] + \
              [frag_avg[metric] for metric in metrics] + [''] + \
              [sub_avg[metric] for metric in metrics]
    sheet.append(avg_row)

    workbook.save(excel_file)

def average_results_and_save_sc_svm(parent_dir: str, numfolds: int, channels: list, svm_act: int, ver: int):
    """
    Goes through all the channel files and averages the results.
    Saves averages into a text file, writes fold results to separate CSV files,
    and appends results to an Excel file with each channel on a separate sheet.
    Metrics (except MCC) are scaled by 100 and presented with 4 significant figures.
    """
    metrics = ['accuracy', 'sensitivity', 'specificity', 'MCC']  # Update these as needed
    scale_metrics = ['accuracy', 'sensitivity', 'specificity']  # Metrics to scale
    excel_dir = os.path.dirname(parent_dir)  # Navigate to the directory above `parent_dir`
    excel_file = os.path.join(excel_dir, 'results_summary_single_svm.xlsx')

    # again = True
    # chan = channels[0]
    for chan in channels:
        frag_sums = {metric: 0 for metric in metrics}
        sub_sums = {metric: 0 for metric in metrics}
        file_count = 0
        fold_results = []  # To store fold-wise results for CSV output

        for fold in range(1, numfolds + 1):
            file = os.path.join(parent_dir, f'ch{chan}', f'fold{fold}', f'svm{svm_act}.txt')
            data = read_json_numpy(file)
            fold_data = {'fold': fold, 'channel': chan, 'svm_act': svm_act}

            # Accumulate values for results_frag, results_sub, and results_sub_soft
            for metric in metrics:
                frag_value = data['results_frag'][metric]
                sub_value = data['results_sub'][metric]

                # Scale metrics if required
                if metric in scale_metrics:
                    frag_value *= 100
                    sub_value *= 100

                # Format to 4 significant figures
                frag_value = round(frag_value, 4)
                sub_value = round(sub_value, 4)

                frag_sums[metric] += frag_value
                sub_sums[metric] += sub_value

                # Add fold data for Excel and CSV
                fold_data[f'frag_{metric}'] = frag_value
                fold_data[f'sub_{metric}'] = sub_value

            fold_results.append(fold_data)
            file_count += 1

        # Compute averages
        frag_avg = {metric: round(frag_sums[metric] / file_count, 4) for metric in metrics}
        sub_avg = {metric: round(sub_sums[metric] / file_count, 4) for metric in metrics}

        # Prepare the output content
        output_data = {
            f'ch{chan}_average_results_frag': frag_avg,
            f'ch{chan}_average_results_sub': sub_avg,
        }

        file_path = parent_dir + '/res_single_svm.txt'
        if os.path.exists(file_path):
            append_json_numpy(output_data, file_path)
        else:
            write_json_numpy(output_data, file_path)

        # Write fold results to CSV
        csv_file = os.path.join(parent_dir, f'ch{chan}_fold_results_svm{svm_act}.csv')
        with open(csv_file, mode='w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=['fold', 'channel','svm_act'] + 
                                    [f'frag_{metric}' for metric in metrics] + [''] + 
                                    [f'sub_{metric}' for metric in metrics])
            
            writer.writeheader()
            writer.writerows(fold_results)

        # Append to Excel file in directory above `parent_dir`
        if os.path.exists(excel_file):
            workbook = load_workbook(excel_file)
        else:
            os.makedirs(excel_dir, exist_ok=True)  # Ensure the directory exists
            workbook = Workbook()

        if f'ch{chan}' not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=f'ch{chan}')
        else:
            sheet = workbook[f'ch{chan}']

        # If the sheet is newly created, write headers
        if sheet.max_row == 1:
            sheet.append(['Version', 'Fold', 'Channel','SVM ACTIVATION'] +
                         [f'frag_{metric}' for metric in metrics] + [''] +
                         [f'sub_{metric}' for metric in metrics])
        
        # Append a version marker and results
        sheet.append([f'Version: {ver}'] + [''] * (len(sheet[1]) - 1))  # Add version marker
        for result in fold_results:
            row = [ver, result['fold'], result['channel'], result['svm_act']] + \
                  [result[f'frag_{metric}'] for metric in metrics] + [''] + \
                  [result[f'sub_{metric}'] for metric in metrics] 
            sheet.append(row)

        # Append averages row after all folds
        avg_row = ['AVERAGE', '',''] + \
                  [''] + [frag_avg[metric] for metric in metrics] + [''] + \
                  [sub_avg[metric] for metric in metrics] + ['']
        sheet.append(avg_row)

        workbook.save(excel_file)

        chan += 1
        if not os.path.exists(os.path.join(parent_dir, f'ch{chan}')):
            again = False

from click import get_current_context
def save_args_to_file(output_dir, script_name="arguments", hidden_options=None):
    """
    Saves the named command-line arguments of the calling script to a .txt file.

    Args:
        output_dir (str): Directory where the arguments file will be saved.
        script_name (str): Name of the script or identifier for the file (default: "arguments").
        hidden_options (list): List of option names to exclude from saving.
    """
    if hidden_options is None:
        hidden_options = []

    # Get the current context and extract parameter values
    ctx = get_current_context()
    args_dict = {key: value for key, value in ctx.params.items() if key not in hidden_options}

    # Ensure the output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Define the output file path
    file_path = os.path.join(output_dir, f"{script_name}_arguments.txt")

    # Save arguments as JSON to a text file
    with open(file_path, 'w') as f:
        json.dump(args_dict, f, indent=4)